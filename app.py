# our mission is to show, how we would modify values be a certain amount for X rows

# importing openpyxl, to work with excel files
# from openpyxl importing the chart module, to create a chart
import openpyxl as xl
from openpyxl.chart import BarChart, Reference

# loading in our data
wb = xl.load_workbook('info.xlsx')
sheet = wb['Sheet1']

# it does not matter how we define our cell
test_cell = sheet['a1']
test_cell1 = sheet.cell(1, 1)
if test_cell == test_cell1:
    print("As we can see, both of our defined cells are equal")
print(f"The value of cell A1 is {test_cell.value}")

# checking how many rows we have
print(f"We have {sheet.max_row} rows")

# adding a for loop, that will multiply the cells in the third column by 0.5

for row in range(2, sheet.max_row+1):
    cell = sheet.cell(row, 3)
    print(f"The value of {cell} is {cell.value}")
    corrected_cell_value = cell.value * 0.5
    corrected_cell = sheet.cell(row, 4)
    corrected_cell.value = corrected_cell_value

# selecting a range of values

values = Reference(sheet,
                   min_row=2,
                   max_row=sheet.max_row,
                   min_col=4,
                   max_col=4)

# creating a chart

chart = BarChart()
chart.add_data(values)

# adding the chart to our sheet

sheet.add_chart(chart, 'e2')

# saving our new file

wb.save('corrected_prices.xlsx')
