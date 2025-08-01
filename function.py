def process_workbook(filename):
    import openpyxl as xl
    from openpyxl.chart import BarChart, Reference
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']
    for row in range(2, sheet.max_row+1):
        cell = sheet.cell(row, 3)
        print(f"The value of {cell} is {cell.value}")
        corrected_cell_value = cell.value * 0.5
        corrected_cell = sheet.cell(row, 4)
        corrected_cell.value = corrected_cell_value
    values = Reference(sheet,
                       min_row=2,
                       max_row=sheet.max_row,
                       min_col=4,
                       max_col=4)
    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')
    wb.save(filename)
